#!/usr/bin/env python3
"""Sync Final Marga Users XLSX to Firestore marga_users collection.

Usage:
  python3 tools/sync-final-marga-users.py "/Users/mike/Downloads/Final Marga Users (1).xlsx"
  python3 tools/sync-final-marga-users.py "/path/file.xlsx" --dry-run
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import json
import os
import re
import secrets
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from typing import Any

import openpyxl

INSECURE_TLS = False


BASE_ROLE_DEFAULTS = {
    "admin": ["customers", "billing", "collections", "service", "inventory", "hr", "reports", "settings", "sync", "field", "purchasing", "pettycash", "sales"],
    "billing": ["customers", "billing", "reports"],
    "collection": ["customers", "collections", "reports"],
    "service": ["customers", "service", "inventory", "field"],
    "hr": ["hr", "settings"],
    "technician": ["field"],
    "messenger": ["field"],
    "viewer": ["customers", "reports"],
}


def parse_firebase_config(path: str) -> tuple[str, str]:
    text = open(path, "r", encoding="utf-8").read()
    api_key_match = re.search(r"apiKey:\s*'([^']+)'", text)
    base_url_match = re.search(r"baseUrl:\s*'([^']+)'", text)
    if not api_key_match or not base_url_match:
        raise RuntimeError("Could not parse apiKey/baseUrl from shared/js/firebase-config.js")
    return api_key_match.group(1), base_url_match.group(1)


def request_json(url: str, method: str = "GET", payload: dict[str, Any] | None = None) -> dict[str, Any] | list[Any]:
    data = None
    headers = {}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    context = ssl._create_unverified_context() if INSECURE_TLS else None
    try:
        with urllib.request.urlopen(req, timeout=60, context=context) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        try:
            parsed = json.loads(raw)
        except Exception:
            parsed = {"error": {"message": raw}}
        message = parsed.get("error", {}).get("message", str(exc))
        raise RuntimeError(message) from exc


def parse_fs_value(value: dict[str, Any]) -> Any:
    if "stringValue" in value:
        return value["stringValue"]
    if "integerValue" in value:
        return int(value["integerValue"])
    if "doubleValue" in value:
        return float(value["doubleValue"])
    if "booleanValue" in value:
        return bool(value["booleanValue"])
    if "timestampValue" in value:
        return value["timestampValue"]
    if "arrayValue" in value:
        return [parse_fs_value(v) for v in value.get("arrayValue", {}).get("values", [])]
    if "mapValue" in value:
        return {k: parse_fs_value(v) for k, v in value.get("mapValue", {}).get("fields", {}).items()}
    return None


def parse_fs_doc(doc: dict[str, Any]) -> dict[str, Any]:
    fields = doc.get("fields", {})
    out = {k: parse_fs_value(v) for k, v in fields.items()}
    name = doc.get("name", "")
    if name:
        out["_docId"] = name.split("/")[-1]
    return out


def to_fs_field(value: Any) -> dict[str, Any]:
    if value is None:
        return {"nullValue": None}
    if isinstance(value, bool):
        return {"booleanValue": value}
    if isinstance(value, int):
        return {"integerValue": str(value)}
    if isinstance(value, float):
        if value.is_integer():
            return {"integerValue": str(int(value))}
        return {"doubleValue": value}
    if isinstance(value, list):
        return {"arrayValue": {"values": [to_fs_field(v) for v in value]}}
    return {"stringValue": str(value)}


def run_query(base_url: str, api_key: str, structured_query: dict[str, Any]) -> list[dict[str, Any]]:
    url = f"{base_url}:runQuery?key={api_key}"
    payload = request_json(url, method="POST", payload={"structuredQuery": structured_query})
    if not isinstance(payload, list):
        return []
    docs = []
    for row in payload:
        doc = row.get("document")
        if doc:
            docs.append(doc)
    return docs


def set_document(base_url: str, api_key: str, collection: str, doc_id: str, fields: dict[str, Any]) -> None:
    fs_fields = {k: to_fs_field(v) for k, v in fields.items()}
    encoded = urllib.parse.quote(doc_id, safe="")
    url = f"{base_url}/{collection}/{encoded}?key={api_key}"
    request_json(url, method="PATCH", payload={"fields": fs_fields})


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def normalize_numeric_text(value: Any) -> str:
    text = str(value if value is not None else "").strip()
    if not text:
        return ""
    if re.match(r"^-?\d+(\.0+)?$", text):
        return str(int(float(text)))
    return text


def normalize_modules(modules: Any) -> list[str]:
    if isinstance(modules, list):
        values = modules
    elif isinstance(modules, str):
        values = modules.split(",")
    else:
        return []
    out = []
    seen = set()
    for item in values:
        value = str(item or "").strip().lower()
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def map_position_to_role(position: str) -> str:
    p = (position or "").strip().lower()
    if not p:
        return "viewer"
    if "admin" in p:
        return "admin"
    if "collect" in p:
        return "collection"
    if any(token in p for token in ("billing", "cashier", "account", "finance", "purchasing")):
        return "billing"
    if "messenger" in p or "driver" in p:
        return "messenger"
    if any(token in p for token in ("tech", "maintenance", "refiller")):
        return "technician"
    if any(token in p for token in ("service", "csr", "sales")):
        return "service"
    if "hr" in p:
        return "hr"
    return "viewer"


def is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email or ""))


def load_role_permissions(base_url: str, api_key: str) -> dict[str, list[str]]:
    role_modules = {role: normalize_modules(mods) for role, mods in BASE_ROLE_DEFAULTS.items()}
    docs = run_query(
        base_url,
        api_key,
        {
            "from": [{"collectionId": "marga_role_permissions"}],
            "orderBy": [{"field": {"fieldPath": "role"}, "direction": "ASCENDING"}],
            "limit": 200,
        },
    )
    for doc in docs:
        parsed = parse_fs_doc(doc)
        role = str(parsed.get("role") or parsed.get("_docId") or "").strip().lower()
        if role not in role_modules:
            continue
        role_modules[role] = normalize_modules(parsed.get("allowed_modules"))
    return role_modules


def build_records(xlsx_path: str, role_modules: dict[str, list[str]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]
    header_idx = -1
    for idx, row in enumerate(rows):
        normalized = [normalize_header(c) for c in row]
        if "employee_id" in normalized and "email" in normalized and "password" in normalized:
            header_idx = idx
            break
    if header_idx < 0:
        raise RuntimeError("Cannot detect XLSX headers. Required columns: Employee_Id, Password, Email.")

    header = [normalize_header(c) for c in rows[header_idx]]
    col = {name: header.index(name) if name in header else -1 for name in ["employee_id", "nickname", "firstname", "lastname", "password", "contact_number", "position", "email"]}
    required = ("employee_id", "password", "email")
    for name in required:
        if col[name] < 0:
            raise RuntimeError(f"Missing required column: {name}")

    records: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        excel_row_num = row_idx + 1
        employee_id_raw = normalize_numeric_text(row[col["employee_id"]])
        nickname = str(row[col["nickname"]] if col["nickname"] >= 0 else "").strip()
        firstname = str(row[col["firstname"]] if col["firstname"] >= 0 else "").strip()
        lastname = str(row[col["lastname"]] if col["lastname"] >= 0 else "").strip()
        password = normalize_numeric_text(row[col["password"]])
        contact_number = normalize_numeric_text(row[col["contact_number"]] if col["contact_number"] >= 0 else "")
        position = str(row[col["position"]] if col["position"] >= 0 else "").strip()
        email = str(row[col["email"]]).strip().lower()

        if not any([employee_id_raw, nickname, firstname, lastname, password, contact_number, position, email]):
            continue
        if not employee_id_raw:
            skipped.append({"row": excel_row_num, "reason": "Missing employee ID"})
            continue
        try:
            staff_id = int(employee_id_raw)
        except Exception:
            skipped.append({"row": excel_row_num, "reason": f"Invalid employee ID: {employee_id_raw}"})
            continue
        if not is_valid_email(email):
            skipped.append({"row": excel_row_num, "reason": f"Invalid/missing email: {email or '(blank)'}"})
            continue
        if not password:
            skipped.append({"row": excel_row_num, "reason": f"Missing password for {email}"})
            continue

        role = map_position_to_role(position)
        modules = normalize_modules(role_modules.get(role, []))
        name = f"{firstname} {lastname}".strip() or nickname or email.split("@")[0]
        records.append(
            {
                "row_number": excel_row_num,
                "staff_id": staff_id,
                "nickname": nickname,
                "firstname": firstname,
                "lastname": lastname,
                "name": name,
                "password": password,
                "contact_number": contact_number,
                "position": position,
                "email": email,
                "role": role,
                "allowed_modules": modules,
            }
        )
    return records, skipped


def hash_password(password: str) -> dict[str, Any]:
    pwd = str(password or "")
    salt = secrets.token_bytes(16)
    iterations = 120000
    derived = hashlib.pbkdf2_hmac("sha256", pwd.encode("utf-8"), salt, iterations, dklen=32)
    return {
        "password_hash": base64.b64encode(derived).decode("ascii"),
        "password_salt": base64.b64encode(salt).decode("ascii"),
        "password_iterations": iterations,
        "password_algo": "PBKDF2-SHA256",
    }


def main() -> int:
    global INSECURE_TLS
    parser = argparse.ArgumentParser(description="Sync Final Marga Users XLSX to Firestore")
    parser.add_argument("xlsx_path", help="Path to Final Marga Users xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Do not write to Firestore")
    parser.add_argument("--insecure", action="store_true", help="Disable TLS certificate verification for this run")
    args = parser.parse_args()
    INSECURE_TLS = bool(args.insecure)

    if not os.path.exists(args.xlsx_path):
        print(f"File not found: {args.xlsx_path}", file=sys.stderr)
        return 2

    api_key, base_url = parse_firebase_config("shared/js/firebase-config.js")
    role_modules = load_role_permissions(base_url, api_key)
    records, skipped = build_records(args.xlsx_path, role_modules)

    print(f"Detected records: {len(records)}")
    print(f"Initial skipped rows: {len(skipped)}")

    synced = 0
    failed: list[dict[str, Any]] = []
    if not args.dry_run:
        for rec in records:
            now = dt.datetime.now(dt.timezone.utc).isoformat()
            fields = {
                "email": rec["email"],
                "username": rec["email"],
                "name": rec["name"],
                "role": rec["role"],
                "active": True,
                "staff_id": rec["staff_id"],
                "allowed_modules": rec["allowed_modules"],
                "allowed_modules_configured": False,
                "nickname": rec["nickname"],
                "firstname": rec["firstname"],
                "lastname": rec["lastname"],
                "position": rec["position"],
                "contact_number": rec["contact_number"],
                "source_file": os.path.basename(args.xlsx_path),
                "source_row": rec["row_number"],
                "imported_at": now,
                "updated_at": now,
            }
            fields.update(hash_password(rec["password"]))
            try:
                set_document(base_url, api_key, "marga_users", rec["email"], fields)
                synced += 1
            except Exception as exc:
                failed.append({"row": rec["row_number"], "email": rec["email"], "reason": str(exc)})
    else:
        synced = len(records)

    all_skipped = skipped + failed
    print(f"Synced: {synced}")
    print(f"Skipped/Failed: {len(all_skipped)}")
    for item in all_skipped[:10]:
        print(f"- row {item['row']}: {item['reason']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
