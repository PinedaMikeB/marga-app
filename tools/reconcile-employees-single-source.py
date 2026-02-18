#!/usr/bin/env python3
"""Reconcile employees into tbl_employee as single source of truth.

Steps:
1) Restore full tbl_employee rows from SQL dump.
2) Mark active/inactive based on Final Marga Users file.
3) Write login fields (email/password hash/role/modules) on tbl_employee docs.
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import json
import os
import re
import secrets
import ssl
import urllib.parse
import urllib.request
from typing import Any

import openpyxl

INSECURE_TLS = False

BASE_ROLE_DEFAULTS = {
    "admin": ["customers", "billing", "collections", "service", "inventory", "hr", "reports", "settings", "sync", "field", "purchasing", "pettycash", "sales"],
    "billing": ["customers", "billing", "reports"],
    "collection": ["customers", "collections", "reports"],
    "service": ["customers", "service", "inventory", "field"],
    "hr": ["hr", "settings"],
    "technician": ["field"],
    "messenger": ["field"],
    "viewer": ["customers", "reports"],
}


def request_json(url: str, method: str = "GET", payload: dict[str, Any] | None = None) -> Any:
    data = None
    headers = {}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    ctx = ssl._create_unverified_context() if INSECURE_TLS else None
    with urllib.request.urlopen(req, timeout=60, context=ctx) as resp:
        raw = resp.read().decode("utf-8")
        return json.loads(raw) if raw else {}


def parse_firebase_config(path: str) -> tuple[str, str]:
    text = open(path, "r", encoding="utf-8").read()
    api_key = re.search(r"apiKey:\s*'([^']+)'", text)
    base_url = re.search(r"baseUrl:\s*'([^']+)'", text)
    if not api_key or not base_url:
        raise RuntimeError("Unable to parse shared/js/firebase-config.js")
    return api_key.group(1), base_url.group(1)


def fs_parse_value(v: dict[str, Any]) -> Any:
    if "stringValue" in v:
        return v["stringValue"]
    if "integerValue" in v:
        return int(v["integerValue"])
    if "doubleValue" in v:
        return float(v["doubleValue"])
    if "booleanValue" in v:
        return bool(v["booleanValue"])
    if "timestampValue" in v:
        return v["timestampValue"]
    if "arrayValue" in v:
        return [fs_parse_value(x) for x in v.get("arrayValue", {}).get("values", [])]
    if "mapValue" in v:
        return {k: fs_parse_value(x) for k, x in v.get("mapValue", {}).get("fields", {}).items()}
    return None


def fs_parse_doc(doc: dict[str, Any]) -> dict[str, Any]:
    out = {k: fs_parse_value(v) for k, v in (doc.get("fields") or {}).items()}
    out["_docId"] = doc.get("name", "").split("/")[-1]
    return out


def fs_field(value: Any) -> dict[str, Any]:
    if value is None:
        return {"nullValue": None}
    if isinstance(value, bool):
        return {"booleanValue": value}
    if isinstance(value, int):
        return {"integerValue": str(value)}
    if isinstance(value, float):
        return {"doubleValue": value}
    if isinstance(value, list):
        return {"arrayValue": {"values": [fs_field(x) for x in value]}}
    if isinstance(value, dict):
        return {"mapValue": {"fields": {k: fs_field(v) for k, v in value.items()}}}
    return {"stringValue": str(value)}


def fetch_collection(base_url: str, api_key: str, collection: str, page_size: int = 1000) -> list[dict[str, Any]]:
    docs: list[dict[str, Any]] = []
    token = ""
    while True:
        q = urllib.parse.urlencode({"pageSize": str(page_size), "key": api_key, **({"pageToken": token} if token else {})})
        payload = request_json(f"{base_url}/{collection}?{q}")
        batch = payload.get("documents") or []
        docs.extend(fs_parse_doc(d) for d in batch)
        token = payload.get("nextPageToken") or ""
        if not token:
            break
    return docs


def set_document(base_url: str, api_key: str, collection: str, doc_id: str, fields: dict[str, Any]) -> None:
    url = f"{base_url}/{collection}/{urllib.parse.quote(str(doc_id), safe='')}?key={api_key}"
    payload = {"fields": {k: fs_field(v) for k, v in fields.items()}}
    request_json(url, method="PATCH", payload=payload)


def parse_mysql_string(raw: str) -> str:
    s = raw[1:-1]
    s = s.replace("\\\\", "\\")
    s = s.replace("\\'", "'")
    s = s.replace("\\r", "\r").replace("\\n", "\n").replace("\\t", "\t")
    return s


def sql_token_to_value(token: str) -> Any:
    t = token.strip()
    if t.upper() == "NULL":
        return None
    if len(t) >= 2 and t[0] == "'" and t[-1] == "'":
        return parse_mysql_string(t)
    if re.fullmatch(r"-?\d+", t):
        return int(t)
    if re.fullmatch(r"-?\d+\.\d+", t):
        return float(t)
    return t


def parse_insert_values(values: str) -> list[list[Any]]:
    rows: list[list[Any]] = []
    row: list[Any] = []
    token: list[str] = []
    in_row = False
    in_str = False
    esc = False

    for ch in values:
        if in_str:
            token.append(ch)
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == "'":
                in_str = False
            continue

        if ch == "'":
            in_str = True
            token.append(ch)
            continue
        if ch == "(":
            in_row = True
            row = []
            token = []
            continue
        if ch == ")" and in_row:
            row.append(sql_token_to_value("".join(token)))
            rows.append(row)
            row = []
            token = []
            in_row = False
            continue
        if ch == "," and in_row:
            row.append(sql_token_to_value("".join(token)))
            token = []
            continue
        if ch == ";" and not in_row:
            break
        if in_row:
            token.append(ch)
    return rows


def extract_tbl_employee_from_dump(dump_path: str) -> tuple[list[str], list[dict[str, Any]]]:
    columns: list[str] = []
    rows: list[dict[str, Any]] = []
    in_create = False
    prefix = "INSERT INTO `tbl_employee` VALUES "

    with open(dump_path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if line.startswith("CREATE TABLE `tbl_employee`"):
                in_create = True
                continue
            if in_create:
                s = line.strip()
                if s.startswith("`"):
                    columns.append(s.split("`")[1])
                    continue
                if s.startswith(") ENGINE="):
                    in_create = False
                    continue
            if line.startswith(prefix):
                value_rows = parse_insert_values(line[len(prefix):])
                for values in value_rows:
                    record = {columns[i]: values[i] if i < len(values) else None for i in range(len(columns))}
                    record["id"] = int(record["id"])
                    rows.append(record)
    if not columns or not rows:
        raise RuntimeError("Failed to parse tbl_employee from dump")
    return columns, rows


def normalize_key(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text or "").strip().lower())


def sanitize_username(text: Any) -> str:
    value = re.sub(r"[^a-z0-9._-]+", "", str(text or "").strip().lower())
    return value.strip("._-")[:48]


def build_username_candidates(rec: dict[str, Any], emp_id: int) -> list[str]:
    out: list[str] = []
    email = str(rec.get("email") or "").strip().lower()
    if email and "@" in email:
        out.append(email.split("@", 1)[0])
    nick = str(rec.get("nickname") or "").strip()
    if nick:
        out.append(nick)
    first = str(rec.get("firstname") or "").strip()
    last = str(rec.get("lastname") or "").strip()
    if first and last:
        out.append(f"{first}.{last}")
    if first:
        out.append(first)
    out.append(f"emp{emp_id}")
    return out


def pick_username(rec: dict[str, Any], emp_id: int, used_usernames: set[str], current_username: str = "") -> str:
    current = sanitize_username(current_username)
    if current:
        used_usernames.discard(current)

    for raw in build_username_candidates(rec, emp_id):
        base = sanitize_username(raw)
        if not base:
            continue
        candidate = base
        i = 2
        while candidate in used_usernames:
            candidate = f"{base}{i}"
            i += 1
        used_usernames.add(candidate)
        return candidate

    fallback = f"emp{emp_id}"
    used_usernames.add(fallback)
    return fallback


def map_position_to_role(position: str) -> str:
    p = str(position or "").strip().lower()
    if "admin" in p or "manager" in p:
        return "admin"
    if "collect" in p:
        return "collection"
    if any(k in p for k in ("billing", "cashier", "account", "finance", "purchasing")):
        return "billing"
    if "messenger" in p or "driver" in p:
        return "messenger"
    if any(k in p for k in ("tech", "maintenance", "refiller")):
        return "technician"
    if any(k in p for k in ("service", "csr", "sales")):
        return "service"
    if "hr" in p:
        return "hr"
    return "viewer"


def parse_final_users_xlsx(path: str) -> list[dict[str, Any]]:
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]
    header_idx = -1
    for i, row in enumerate(rows):
        norm = [normalize_key(v) for v in row]
        if "employeeid" in norm and "firstname" in norm and "lastname" in norm:
            header_idx = i
            break
    if header_idx < 0:
        raise RuntimeError("Cannot detect header row in Final Marga Users xlsx")

    header = [normalize_key(v) for v in rows[header_idx]]
    idx = {name: header.index(name) if name in header else -1 for name in ["employeeid", "nickname", "firstname", "lastname", "password", "contactnumber", "position", "email"]}
    out: list[dict[str, Any]] = []
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        first = str(row[idx["firstname"]] if idx["firstname"] >= 0 else "").strip()
        last = str(row[idx["lastname"]] if idx["lastname"] >= 0 else "").strip()
        nick = str(row[idx["nickname"]] if idx["nickname"] >= 0 else "").strip()
        email = str(row[idx["email"]] if idx["email"] >= 0 else "").strip().lower()
        if not (first or last or nick or email):
            continue
        password = str(row[idx["password"]] if idx["password"] >= 0 else "").strip()
        if re.fullmatch(r"-?\d+(\.0+)?", password):
            password = str(int(float(password)))
        rec = {
            "row": r + 1,
            "employee_id": int(float(row[idx["employeeid"]])) if idx["employeeid"] >= 0 and str(row[idx["employeeid"]]).strip() else None,
            "nickname": nick,
            "firstname": first,
            "lastname": last,
            "full_name_key": f"{normalize_key(first)}|{normalize_key(last)}",
            "nick_last_key": f"{normalize_key(nick)}|{normalize_key(last)}",
            "email": email,
            "email_valid": bool(re.fullmatch(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email or "")),
            "password": password,
            "has_password": bool(password),
            "position": str(row[idx["position"]] if idx["position"] >= 0 else "").strip(),
            "contact_number": str(row[idx["contactnumber"]] if idx["contactnumber"] >= 0 else "").strip(),
        }
        out.append(rec)
    return out


def hash_password(password: str) -> dict[str, Any]:
    salt = secrets.token_bytes(16)
    iterations = 120000
    derived = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations, dklen=32)
    return {
        "password_hash": base64.b64encode(derived).decode("ascii"),
        "password_salt": base64.b64encode(salt).decode("ascii"),
        "password_iterations": iterations,
        "password_algo": "PBKDF2-SHA256",
    }


def main() -> int:
    global INSECURE_TLS
    parser = argparse.ArgumentParser(description="Reconcile tbl_employee as single source")
    parser.add_argument("--dump", default="/Users/mike/Downloads/Dump20260218.sql")
    parser.add_argument("--xlsx", default="/Users/mike/Downloads/Final Marga Users (1).xlsx")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--insecure", action="store_true")
    args = parser.parse_args()
    INSECURE_TLS = args.insecure

    api_key, base_url = parse_firebase_config("shared/js/firebase-config.js")
    _, dump_rows = extract_tbl_employee_from_dump(args.dump)
    existing_docs = fetch_collection(base_url, api_key, "tbl_employee", 1000)
    existing_by_id = {int(d["id"]): d for d in existing_docs if isinstance(d.get("id"), int)}
    role_modules = BASE_ROLE_DEFAULTS.copy()
    for doc in fetch_collection(base_url, api_key, "marga_role_permissions", 200):
        role = str(doc.get("role") or doc.get("_docId") or "").strip().lower()
        if role in role_modules:
            role_modules[role] = [str(x).strip().lower() for x in (doc.get("allowed_modules") or []) if str(x).strip()]

    docs_by_id: dict[int, dict[str, Any]] = {}
    for row in dump_rows:
        rid = int(row["id"])
        merged = {}
        if rid in existing_by_id:
            merged.update(existing_by_id[rid])
        merged.update(row)
        merged["id"] = rid
        merged["marga_active"] = False
        merged["marga_account_active"] = False
        current_role = str(merged.get("marga_role") or "").strip().lower() or "viewer"
        if current_role not in BASE_ROLE_DEFAULTS:
            current_role = "viewer"
        merged["marga_role"] = current_role
        merged["marga_allowed_modules"] = role_modules.get(current_role, [])
        merged["allowed_modules_configured"] = False
        docs_by_id[rid] = merged

    # Keep every existing numeric-id employee doc in the unified table, but force
    # records outside the SQL dump inactive so old/new duplicate sources cannot stay active.
    dump_ids = set(docs_by_id.keys())
    for existing in existing_docs:
        rid = existing.get("id")
        if not isinstance(rid, int) or rid in dump_ids:
            continue
        merged = dict(existing)
        merged["id"] = rid
        merged["marga_active"] = False
        merged["marga_account_active"] = False
        current_role = str(merged.get("marga_role") or "").strip().lower() or "viewer"
        if current_role not in BASE_ROLE_DEFAULTS:
            current_role = "viewer"
        merged["marga_role"] = current_role
        merged["marga_allowed_modules"] = role_modules.get(current_role, [])
        merged["allowed_modules_configured"] = False
        docs_by_id[rid] = merged

    final_rows = parse_final_users_xlsx(args.xlsx)
    by_first_last: dict[str, list[int]] = {}
    by_nick_last: dict[str, list[int]] = {}
    for rid, employee in docs_by_id.items():
        first_last = f"{normalize_key(employee.get('firstname'))}|{normalize_key(employee.get('lastname'))}"
        nick_last = f"{normalize_key(employee.get('nickname'))}|{normalize_key(employee.get('lastname'))}"
        by_first_last.setdefault(first_last, []).append(rid)
        by_nick_last.setdefault(nick_last, []).append(rid)

    matched = 0
    unmatched: list[dict[str, Any]] = []
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    matched_ids: set[int] = set()
    used_usernames: set[str] = set()
    for employee in docs_by_id.values():
        existing_username = sanitize_username(employee.get("username"))
        if existing_username:
            used_usernames.add(existing_username)

    for rec in final_rows:
        candidates: list[int] = []
        if rec["employee_id"] in docs_by_id:
            candidates.append(rec["employee_id"])
        candidates += by_first_last.get(rec["full_name_key"], [])
        candidates += by_nick_last.get(rec["nick_last_key"], [])
        candidates = list(dict.fromkeys(candidates))
        candidates = [cid for cid in candidates if cid not in matched_ids] or candidates
        if not candidates:
            unmatched.append({"row": rec["row"], "name": f"{rec['firstname']} {rec['lastname']}".strip(), "reason": "no employee match"})
            continue
        emp_id = candidates[0]
        if len(candidates) > 1:
            active_first = [cid for cid in candidates if int(docs_by_id[cid].get("estatus") or 0) == 1]
            if active_first:
                emp_id = active_first[0]
        emp = docs_by_id[emp_id]

        role = map_position_to_role(rec["position"])
        emp["marga_active"] = True
        emp["marga_account_active"] = True
        emp["marga_role"] = role
        emp["marga_allowed_modules"] = role_modules.get(role, [])
        emp["allowed_modules_configured"] = False
        emp["marga_updated_at"] = now
        emp["marga_source_file"] = os.path.basename(args.xlsx)
        emp["marga_source_row"] = rec["row"]
        if rec["contact_number"]:
            emp["contact_number"] = rec["contact_number"]
        if rec["email_valid"]:
            emp["email"] = rec["email"]
            emp["marga_login_email"] = rec["email"]
        username = pick_username(rec, emp_id, used_usernames, str(emp.get("username") or ""))
        if username:
            emp["username"] = username
        if rec["has_password"]:
            emp.update(hash_password(rec["password"]))
            emp["marga_password_updated_at"] = now

        matched += 1
        matched_ids.add(emp_id)

    active_count = sum(1 for d in docs_by_id.values() if d.get("marga_active") is True)
    inactive_count = len(docs_by_id) - active_count
    print(f"Dump employees: {len(dump_rows)}")
    print(f"Final user rows: {len(final_rows)}")
    print(f"Matched active: {matched}")
    print(f"Unmatched rows: {len(unmatched)}")
    print(f"Result active: {active_count}, inactive: {inactive_count}")
    if unmatched:
        for row in unmatched[:10]:
            print(f"- row {row['row']}: {row['name']} ({row['reason']})")

    if args.dry_run:
        return 0

    for rid in sorted(docs_by_id):
        set_document(base_url, api_key, "tbl_employee", str(rid), docs_by_id[rid])
    print(f"Wrote {len(docs_by_id)} employee docs to tbl_employee")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
